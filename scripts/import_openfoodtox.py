#!/usr/bin/env python3
"""Build a local searchable OpenFoodTox SQLite index from EFSA exports.

This is the non-IUCLID path for regular KEvidence installations. Download the
OpenFoodTox Excel export from EFSA/Zenodo, then run for example:

    python scripts/import_openfoodtox.py --download-latest --db data/openfoodtox.db

The importer intentionally stores source rows generically because OpenFoodTox
sheet names and column names may evolve. The application classifies matching
rows at query time into substances, toxicological values/study records, and
publication/source-link records.
"""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
import urllib.error
import urllib.request
from pathlib import Path
from typing import Iterable, Any


def clean_cell(value):
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def iter_xlsx_rows(path: Path) -> Iterable[tuple[str, dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx imports. Install requirements.txt first.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = None
        for raw in rows:
            values = [clean_cell(v) for v in raw]
            if not any(values):
                continue
            if headers is None:
                headers = [v or f"column_{i + 1}" for i, v in enumerate(values)]
                continue
            row = {headers[i] if i < len(headers) else f"column_{i + 1}": values[i] for i in range(len(values))}
            if any(row.values()):
                yield sheet.title, row


def iter_delimited_rows(path: Path) -> Iterable[tuple[str, dict[str, str]]]:
    delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        for row in reader:
            cleaned = {clean_cell(k): clean_cell(v) for k, v in row.items() if k is not None}
            if any(cleaned.values()):
                yield path.stem, cleaned


def iter_input_rows(path: Path) -> Iterable[tuple[str, dict[str, str]]]:
    if path.is_dir():
        for child in sorted(path.iterdir()):
            if child.suffix.lower() in {".csv", ".tsv", ".tab", ".xlsx", ".xlsm"}:
                yield from iter_input_rows(child)
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        yield from iter_xlsx_rows(path)
    elif path.suffix.lower() in {".csv", ".tsv", ".tab"}:
        yield from iter_delimited_rows(path)
    else:
        raise SystemExit(f"Unsupported input format: {path}")


ZENODO_RECORD_ID = "19388272"
ZENODO_API_URL = f"https://zenodo.org/api/records/{ZENODO_RECORD_ID}"


def download_latest_openfoodtox_excel(cache_dir: Path) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    try:
        with urllib.request.urlopen(ZENODO_API_URL, timeout=60) as response:
            record: dict[str, Any] = json.loads(response.read().decode("utf-8"))
    except urllib.error.URLError as exc:
        raise SystemExit(f"Could not reach Zenodo to download OpenFoodTox: {exc}. If this server has restricted outbound access, download the Excel file manually and rerun with --input /path/to/file.xlsx.") from exc

    files = record.get("files") or []
    xlsx_file = next((f for f in files if str(f.get("key", "")).lower().endswith(".xlsx")), None)
    if not xlsx_file:
        raise SystemExit("No .xlsx file found in the OpenFoodTox Zenodo record.")

    filename = Path(xlsx_file["key"]).name
    output_path = cache_dir / filename
    links = xlsx_file.get("links") or {}
    download_url = links.get("self") or links.get("download")
    if not download_url:
        raise SystemExit("The OpenFoodTox Zenodo record did not include a download URL for the Excel file.")

    print(f"Downloading {filename} from Zenodo record {ZENODO_RECORD_ID}...")
    try:
        with urllib.request.urlopen(download_url, timeout=300) as response, output_path.open("wb") as out:
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                out.write(chunk)
    except urllib.error.URLError as exc:
        raise SystemExit(f"Could not download {filename} from Zenodo: {exc}. If this server has restricted outbound access, download the Excel file manually and rerun with --input /path/to/file.xlsx.") from exc
    return output_path


def build_index(input_path: Path, db_path: Path) -> int:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS oft_rows")
    cur.execute("DROP TABLE IF EXISTS oft_metadata")
    cur.execute("CREATE TABLE oft_rows (id INTEGER PRIMARY KEY, table_name TEXT, row_text TEXT, row_json TEXT)")
    cur.execute("CREATE INDEX idx_oft_rows_table ON oft_rows(table_name)")
    cur.execute("CREATE TABLE oft_metadata (key TEXT PRIMARY KEY, value TEXT)")

    count = 0
    for table_name, row in iter_input_rows(input_path):
        row_json = json.dumps(row, ensure_ascii=False, sort_keys=True)
        row_text = f"{table_name} " + " ".join(str(v) for v in row.values())
        cur.execute(
            "INSERT INTO oft_rows (table_name, row_text, row_json) VALUES (?, ?, ?)",
            (table_name, row_text.lower(), row_json),
        )
        count += 1
        if count % 5000 == 0:
            conn.commit()

    cur.execute("INSERT INTO oft_metadata (key, value) VALUES (?, ?)", ("source_path", str(input_path)))
    cur.execute("INSERT INTO oft_metadata (key, value) VALUES (?, ?)", ("row_count", str(count)))
    conn.commit()
    conn.close()
    return count


def main():
    parser = argparse.ArgumentParser(description="Import EFSA OpenFoodTox Excel/CSV exports into a KEvidence SQLite index.")
    parser.add_argument("--input", help="OpenFoodTox .xlsx file, .csv/.tsv file, or directory of exported sheets")
    parser.add_argument("--download-latest", action="store_true", help="Download the latest OpenFoodTox Excel export from the official Zenodo record before importing")
    parser.add_argument("--cache-dir", default="data/openfoodtox_downloads", help="Where to store the downloaded Excel file")
    parser.add_argument("--db", default="data/openfoodtox.db", help="Output SQLite DB path")
    args = parser.parse_args()

    if args.download_latest:
        input_path = download_latest_openfoodtox_excel(Path(args.cache_dir))
    elif args.input:
        input_path = Path(args.input)
    else:
        raise SystemExit("Provide --input /path/to/export.xlsx or use --download-latest")

    count = build_index(input_path, Path(args.db))
    print(f"Imported {count} OpenFoodTox rows into {args.db}")


if __name__ == "__main__":
    main()
